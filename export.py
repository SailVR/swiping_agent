"""
数据导出模块

支持将查询结果导出为：
- Excel (.xlsx) — 表格数据，带样式表头
- PDF (.pdf) — 回答文本 + 数据明细表

依赖：pip install openpyxl fpdf2
"""

import json
import io
import os
from typing import Optional, List, Dict, Any

from logger import get_logger

logger = get_logger(__name__)

# ─── PDF 中文字体路径检测 ────────────────────────────────

_CHINESE_FONTS = [
    # macOS
    "/System/Library/Fonts/PingFang.ttc",
    "/System/Library/Fonts/STHeiti Light.ttc",
    "/System/Library/Fonts/Supplemental/Songti.ttc",
    # Windows
    "C:/Windows/Fonts/msyh.ttc",       # Microsoft YaHei
    "C:/Windows/Fonts/simsun.ttc",      # SimSun
    "C:/Windows/Fonts/simhei.ttf",      # SimHei
    # Linux (noto)
    "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/noto-cjk/NotoSansCJK-Regular.ttc",
]


def _find_chinese_font() -> Optional[str]:
    """查找系统中可用的中文字体路径。"""
    for fp in _CHINESE_FONTS:
        if os.path.exists(fp):
            return fp
    return None


# ─── 数据解析工具 ────────────────────────────────────────


def _parse_data(data_json: str) -> List[Dict[str, Any]]:
    """将后端存储的 JSON 字符串解析为行数据列表。"""
    if not data_json or data_json.strip() in ("{}", '{"message": "查询结果为空"}'):
        return []
    try:
        data = json.loads(data_json)
    except (json.JSONDecodeError, TypeError):
        return []
    if isinstance(data, dict):
        return [data] if "message" not in data else []
    if isinstance(data, list):
        return data
    return []


def _get_columns(data: List[Dict]) -> List[str]:
    """从数据第一行提取列名。"""
    return list(data[0].keys()) if data else []


def _clean_text(text: str) -> str:
    """清理导出的文本（替换 emoji 为文字描述，移除无法渲染的字符）。"""
    emoji_map = {
        "💰": "$ ", "📊": "[Chart] ", "🌐": "[Web] ", "🔍": "[Search] ",
        "💡": "[Tip] ", "✅": "[Yes] ", "❌": "[No] ", "⚠️": "[Warning] ",
        "🎉": "[Great] ", "📈": "[Up] ", "📉": "[Down] ", "🏆": "[Top] ",
        "👋": "[Hi] ", "💬": "[Chat] ", "🔁": "[Retry] ", "🧠": "[Memory] ",
        "📁": "[File] ", "📝": "[Note] ",
    }
    for emoji, repl in emoji_map.items():
        text = text.replace(emoji, repl)
    # 移除无法渲染的辅助字符（非 BMP）
    text = "".join(c for c in text if ord(c) < 0x10000 or c in "\n\r\t")
    return text


# ═══════════════════════════════════════════════════════════
#  Excel 导出
# ═══════════════════════════════════════════════════════════


def _echarts_to_openpyxl_chart(ws, chart_config: dict, columns: list,
                               row_count: int) -> None:
    """将 ECharts 图表配置转换为 openpyxl 原生图表，插入到数据表下方。

    支持 bar / line / pie 三种图表类型，自动映射到 openpyxl 的
    BarChart / LineChart / PieChart。

    Args:
        ws: openpyxl Worksheet 对象。
        chart_config: ECharts 配置字典（含 series / title / xAxis 等）。
        columns: 数据列名列表（用于定位数据源）。
        row_count: 数据行数（不含表头）。
    """
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList

    series_list = chart_config.get("series", [])
    if not series_list or len(columns) < 2:
        return

    echarts_type = series_list[0].get("type", "bar")
    chart_title = chart_config.get("title", {}).get("text", "")

    # ── 映射图表类型 ──
    chart_map = {"bar": BarChart, "line": LineChart, "pie": PieChart}
    ChartCls = chart_map.get(echarts_type)
    if ChartCls is None:
        logger.warning("不支持的图表类型: %s，跳过图表导出", echarts_type)
        return

    # ── 确定分类列和数值列 ──
    # 第一个文本列作为分类轴，第一个数值列作为数值轴
    cat_col_idx = 1   # 默认第一列
    val_col_idx = 2   # 默认第二列

    # 对于饼图，需要 label 和 value
    if echarts_type == "bar":
        chart = BarChart()
        chart.type = "col"  # 柱状图（纵向）
    elif echarts_type == "line":
        chart = LineChart()
        chart.smooth = False
    else:
        chart = PieChart()

    chart.title = chart_title or "图表"
    chart.style = 10

    # ── 数据引用 ──
    # 分类轴：第一列
    cats = Reference(ws, min_col=cat_col_idx, min_row=2,
                     max_row=row_count + 1)

    # 数值：第二列
    data_ref = Reference(ws, min_col=val_col_idx, min_row=1,
                         max_row=row_count + 1)

    if echarts_type == "pie":
        chart.add_data(data_ref)
        chart.set_categories(cats)
        # 百分比 + 分类名标签
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True
    else:
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.x_axis.title = columns[cat_col_idx - 1]
        chart.y_axis.title = columns[val_col_idx - 1]

    # 如果有多个系列，遍历添加
    for i in range(1, len(series_list)):
        extra_col = val_col_idx + i
        if extra_col > len(columns):
            break
        extra_ref = Reference(ws, min_col=extra_col, min_row=1,
                              max_row=row_count + 1)
        chart.add_data(extra_ref, titles_from_data=True)

    # ── 定位到数据下方（留 2 行间距） ──
    chart_anchor_row = row_count + 3
    chart.anchor = f"A{chart_anchor_row}"
    chart.width = 22
    chart.height = 14

    ws.add_chart(chart)


def export_to_excel(data_json: str, title: str = "查询结果",
                    chart_config: Optional[dict] = None) -> io.BytesIO:
    """将查询结果 JSON 导出为格式化的 Excel (.xlsx)。

    Args:
        data_json: SQL 查询结果 JSON 字符串。
        title: 工作表标签名。

    Returns:
        包含 .xlsx 内容的 BytesIO 对象。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    rows = _parse_data(data_json)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel 工作表名最长 31 字符

    if not rows:
        ws.cell(row=1, column=1, value="（查询结果为空）")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    columns = _get_columns(rows)

    # ── 样式定义 ──
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Microsoft YaHei")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ── 写表头 ──
    for ci, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── 写数据行 ──
    for ri, row in enumerate(rows, 2):
        for ci, col_name in enumerate(columns, 1):
            val = row.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = cell_align

    # ── 自动列宽（上限 50 字符） ──
    for ci, col_name in enumerate(columns, 1):
        col_letter = get_column_letter(ci)
        max_len = len(str(col_name))
        for ri in range(2, len(rows) + 2):
            cell_val = str(ws.cell(row=ri, column=ci).value or "")
            max_len = max(max_len, len(cell_val))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # ── 插入图表（如果有 ECharts 配置） ──
    if chart_config and rows:
        try:
            _echarts_to_openpyxl_chart(ws, chart_config, columns, len(rows))
        except Exception as e:
            logger.warning("图表导出失败（不影响数据导出）: %s", e)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════
#  PDF 导出
# ═══════════════════════════════════════════════════════════


def export_to_pdf(answer_text: str, data_json: Optional[str] = None,
                  title: str = "Query Report") -> io.BytesIO:
    """将回答文本和数据导出为 PDF 文件。

    会自动检测系统已安装的中文字体；找不到时使用 Helvetica（仅 ASCII）。

    Args:
        answer_text: 完整的回答文本（支持换行）。
        data_json: 可选的 SQL 查询结果 JSON。
        title: PDF 标题。

    Returns:
        包含 PDF 内容的 BytesIO 对象。
    """
    from fpdf import FPDF

    # ── 自定义 PDF 类 ──
    class ReportPDF(FPDF):
        def header(self):
            if self.page_no() > 1:
                self.set_font("Helvetica", "I", 8)
                self.cell(0, 8, title, align="C", new_x="LMARGIN", new_y="NEXT")
                self.line(10, 18, 200, 18)
                self.ln(4)

        def footer(self):
            self.set_y(-15)
            self.set_font("Helvetica", "I", 8)
            self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")

    pdf = ReportPDF()
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)

    # ── 尝试加载中文字体 ──
    font_path = _find_chinese_font()
    has_cjk = font_path is not None

    def set_body_font(style="", size=10):
        if has_cjk:
            # 把同一个字体注册到所有字重（B/I/BI），防止 set_font(..., "B") 报 Undefined font
            pdf.add_font("CJK", "", font_path)
            pdf.add_font("CJK", "B", font_path)
            pdf.add_font("CJK", "I", font_path)
            pdf.add_font("CJK", "BI", font_path)
            pdf.set_font("CJK", style, size)
        else:
            pdf.set_font("Helvetica", style, size)
            logger.info("未找到中文字体，PDF 中文内容将无法显示。")

    pdf.add_page()

    # ── 标题 ──
    set_body_font("B", 18)
    pdf.cell(0, 15, _clean_text(title), align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)

    # ── 回答正文 ──
    set_body_font(size=11)
    for line in answer_text.split("\n"):
        line = _clean_text(line.strip())
        if not line:
            pdf.ln(4)
            continue
        # multi_cell 自动换行
        pdf.multi_cell(0, 6, line)

    # ── 数据明细表 ──
    if data_json:
        data_rows = _parse_data(data_json)
        if data_rows:
            pdf.add_page()
            set_body_font("B", 14)
            pdf.cell(0, 10, "Data Details", align="C", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(6)

            columns = _get_columns(data_rows)
            col_count = len(columns)

            if col_count == 0:
                pass  # 无列则跳过表格
            else:
                # 可用页面宽度（A4: 210mm, 左右各 10mm 边距 → 190mm）
                page_w = pdf.w - pdf.l_margin - pdf.r_margin
                col_w = page_w / col_count  # 平均分配，不设硬下限

                # 根据列宽自动选择字号
                if col_w < 16:
                    head_size, body_size = 6, 5
                elif col_w < 24:
                    head_size, body_size = 7, 6
                else:
                    head_size, body_size = 8, 7

                def _truncate_to_fit(text: str, width: float, font_style: str = "",
                                     font_size: int = 8) -> str:
                    """将文本截断到指定宽度（用 get_string_width 精确测量）。"""
                    set_body_font(font_style, font_size)
                    clean = _clean_text(str(text))
                    if not clean:
                        return ""
                    while clean and pdf.get_string_width(clean) > width - 1:
                        clean = clean[:-1]
                    return clean

                # ── 表头行 ──
                for col in columns:
                    display = _truncate_to_fit(col, col_w, "B", head_size)
                    pdf.cell(col_w, 7, display, border=1, align="C")
                pdf.ln()

                # ── 数据行（最多 100 行） ──
                for row in data_rows[:100]:
                    for col in columns:
                        val = row.get(col, "")
                        display = _truncate_to_fit(val, col_w, "", body_size)
                        pdf.cell(col_w, 6, display, border=1, align="C")
                    pdf.ln()

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf
